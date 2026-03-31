import csv
import io
import uuid
from datetime import datetime
from decimal import Decimal

from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.conversation import Conversation, Message
from app.models.employee import Employee
from app.models.metrics import ResponseTimeMetric
from app.models.satisfaction import SatisfactionSurvey
from app.models.student import Student
from app.models.ticket import Ticket


async def get_conversation_report(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    start_date: datetime,
    end_date: datetime,
) -> list[dict]:
    result = await db.execute(
        select(Conversation)
        .where(
            Conversation.tenant_id == tenant_id,
            Conversation.started_at >= start_date,
            Conversation.started_at <= end_date,
        )
        .order_by(Conversation.started_at.desc())
    )
    conversations = result.scalars().all()

    rows = []
    for conv in conversations:
        rows.append(
            {
                "id": str(conv.id),
                "status": conv.status,
                "context_type": conv.context_type or "",
                "started_at": conv.started_at.isoformat(),
                "last_message_at": conv.last_message_at.isoformat() if conv.last_message_at else "",
                "satisfaction_score": conv.satisfaction_score or "",
            }
        )
    return rows


async def get_ticket_report(
    db: AsyncSession,
    tenant_id: uuid.UUID,
    start_date: datetime,
    end_date: datetime,
) -> list[dict]:
    result = await db.execute(
        select(Ticket)
        .where(
            Ticket.tenant_id == tenant_id,
            Ticket.created_at >= start_date,
            Ticket.created_at <= end_date,
        )
        .order_by(Ticket.created_at.desc())
    )
    tickets = result.scalars().all()

    rows = []
    for t in tickets:
        rows.append(
            {
                "protocol": t.protocol_number,
                "subject": t.subject,
                "category": t.category or "",
                "priority": t.priority,
                "status": t.status,
                "created_at": t.created_at.isoformat(),
                "sla_deadline": t.sla_deadline.isoformat() if t.sla_deadline else "",
                "resolved_at": t.resolved_at.isoformat() if t.resolved_at else "",
            }
        )
    return rows


def rows_to_csv(rows: list[dict]) -> bytes:
    """Converte lista de dicts para CSV em bytes."""
    if not rows:
        return b""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=rows[0].keys())
    writer.writeheader()
    writer.writerows(rows)
    return output.getvalue().encode("utf-8-sig")  # BOM para Excel


def rows_to_pdf(rows: list[dict], title: str = "Relatório IGS") -> bytes:
    """Converte lista de dicts para PDF em bytes usando ReportLab."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    if not rows:
        return b""

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1 * cm, bottomMargin=1 * cm)
    styles = getSampleStyleSheet()
    elements = []

    elements.append(Paragraph(title, styles["Heading1"]))
    elements.append(Spacer(1, 0.5 * cm))

    headers = list(rows[0].keys())
    table_data = [headers]
    for row in rows:
        table_data.append([str(v) for v in row.values()])

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B3A5C")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
                ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ]
        )
    )
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


def rows_to_excel(rows: list[dict], sheet_name: str = "Relatório") -> bytes:
    """Converte lista de dicts para Excel (.xlsx) usando openpyxl."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    if not rows:
        return b""

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    headers = list(rows[0].keys())
    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    for row_idx, row in enumerate(rows, 2):
        for col_idx, key in enumerate(headers, 1):
            value = row[key]
            if isinstance(value, Decimal):
                value = float(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = 0
        column_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[column_letter].width = min(max_len + 2, 40)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def get_dashboard_overview(db: AsyncSession, tenant_id: uuid.UUID) -> dict:
    from datetime import timedelta, timezone

    now = datetime.now(timezone.utc)
    today_start = now.replace(hour=0, minute=0, second=0, microsecond=0)
    week_start = today_start - timedelta(days=7)
    month_start = today_start - timedelta(days=30)

    # Conversas hoje
    r1 = await db.execute(
        select(func.count(Conversation.id)).where(
            Conversation.tenant_id == tenant_id,
            Conversation.started_at >= today_start,
        )
    )
    today_convs = r1.scalar() or 0

    # Conversas semana
    r2 = await db.execute(
        select(func.count(Conversation.id)).where(
            Conversation.tenant_id == tenant_id,
            Conversation.started_at >= week_start,
        )
    )
    week_convs = r2.scalar() or 0

    # Conversas mês
    r3 = await db.execute(
        select(func.count(Conversation.id)).where(
            Conversation.tenant_id == tenant_id,
            Conversation.started_at >= month_start,
        )
    )
    month_convs = r3.scalar() or 0

    # Tickets abertos
    r4 = await db.execute(
        select(func.count(Ticket.id)).where(
            Ticket.tenant_id == tenant_id,
            Ticket.status.in_(["open", "in_progress"]),
        )
    )
    open_tickets = r4.scalar() or 0

    # Tickets com SLA violado
    r5 = await db.execute(
        select(func.count(Ticket.id)).where(
            Ticket.tenant_id == tenant_id,
            Ticket.status.in_(["open", "in_progress"]),
            Ticket.sla_deadline < now,
        )
    )
    sla_breached = r5.scalar() or 0

    # Taxa de resolução automática (mensagens com sender_type=bot / total)
    r6 = await db.execute(
        select(func.count(Conversation.id)).where(
            Conversation.tenant_id == tenant_id,
            Conversation.status == "closed",
            Conversation.started_at >= month_start,
        )
    )
    closed = r6.scalar() or 0

    r7 = await db.execute(
        select(func.count(Conversation.id)).where(
            Conversation.tenant_id == tenant_id,
            Conversation.status == "closed",
            Conversation.assigned_agent_id.is_(None),
            Conversation.started_at >= month_start,
        )
    )
    auto_closed = r7.scalar() or 0

    auto_rate = (auto_closed / closed * 100) if closed > 0 else 0.0

    # Tempo médio de resposta (do banco)
    avg_rt = await db.scalar(
        select(func.avg(ResponseTimeMetric.response_time_seconds)).where(
            ResponseTimeMetric.tenant_id == tenant_id,
            ResponseTimeMetric.created_at >= month_start,
        )
    )

    # Satisfação média
    avg_sat = await db.scalar(
        select(func.avg(SatisfactionSurvey.score)).where(
            SatisfactionSurvey.tenant_id == tenant_id,
            SatisfactionSurvey.created_at >= month_start,
        )
    )

    # Total alunos e funcionários
    total_students = (
        await db.scalar(select(func.count(Student.id)).where(Student.tenant_id == tenant_id)) or 0
    )
    total_employees = (
        await db.scalar(select(func.count(Employee.id)).where(Employee.tenant_id == tenant_id)) or 0
    )

    # Mensagens e tokens no mês
    total_messages = (
        await db.scalar(
            select(func.count(Message.id)).where(
                Message.tenant_id == tenant_id,
                Message.created_at >= month_start,
            )
        )
        or 0
    )
    ai_tokens = (
        await db.scalar(
            select(func.coalesce(func.sum(Message.ai_tokens_used), 0)).where(
                Message.tenant_id == tenant_id,
                Message.created_at >= month_start,
            )
        )
        or 0
    )

    # Economia estimada: cada conversa auto-resolvida evita ~R$8 de custo humano
    cost_savings = round(auto_closed * 8.0, 2)

    return {
        "total_conversations_today": today_convs,
        "total_conversations_week": week_convs,
        "total_conversations_month": month_convs,
        "auto_resolution_rate": round(auto_rate, 1),
        "avg_response_time_seconds": round(float(avg_rt or 0), 2),
        "open_tickets": open_tickets,
        "sla_breached_tickets": sla_breached,
        "active_agents": 0,
        "avg_satisfaction_score": round(float(avg_sat or 0), 2),
        "total_students": total_students,
        "total_employees": total_employees,
        "total_messages_month": total_messages,
        "ai_tokens_month": ai_tokens,
        "estimated_cost_savings": cost_savings,
    }
